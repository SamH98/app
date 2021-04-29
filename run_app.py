from flask import Flask, render_template, url_for, request, redirect, make_response
import pandas as pd
import pyshark
from openpyxl import Workbook
import scapy.all as scapy
from sklearn.ensemble import IsolationForest
import time
from sklearn.preprocessing import StandardScaler, LabelEncoder

time = 15
runningtime = 0

app = Flask(__name__)
book = Workbook()
sheet = book.active
book1 = Workbook()
sheet1 = book1.active
book2 = Workbook()
sheet2 = book2.active
book3 = Workbook()
sheet3 = book3.active

def setuphosts(runningtime):
    if runningtime == 0:
        sheet1['A1'] = 'IP ADDRESS'
        sheet1['B1'] = 'MAC ADDRESS'
        sheet1['C1'] = 'Device name'
        sheet1['D1'] = 'Active status'
        book1.save('devices.xlsx')

        request = scapy.ARP()
        request.pdst = '192.168.0.1/24'
        broadcast = scapy.Ether()
        broadcast.dst = 'ff:ff:ff:ff:ff:ff'
        request_broadcast = broadcast / request
        clients = scapy.srp(request_broadcast, timeout=1)[0]

        for element in clients:
            device = (element[1].psrc, element[1].hwsrc, 'Unidentified', 'Active')
            sheet1.append(device)
            book1.save('devices.xlsx')
        runningtime += 1
    else:
        return


def refreshhosts(alert):
    request = scapy.ARP()
    request.pdst = '192.168.0.1/24'
    broadcast = scapy.Ether()
    broadcast.dst = 'ff:ff:ff:ff:ff:ff'
    request_broadcast = broadcast / request
    clients = scapy.srp(request_broadcast, timeout=1)[0]
    activedevices = pd.read_excel('devices.xlsx')
    founddevices = pd.DataFrame(columns=['IP ADDRESS', 'MAC ADDRESS', 'Device name', 'Active status'], )

    for element in clients:
        device = ([element[1].psrc, element[1].hwsrc, 'Unidentified', 'Active'])

        if element[1].psrc not in activedevices.values:
            df_length = len(activedevices)
            activedevices.loc[df_length] = device
            df_length = len(founddevices)
            founddevices.loc[df_length] = device
            alert = "New device found IP:" + element[1].psrc
        else:
            df_length = len(founddevices)
            founddevices.loc[df_length] = device

    device_diff = pd.concat([activedevices, founddevices], ).drop_duplicates(keep=False)
    device_diff['Active status'] = device_diff['Active status'].replace(['Active'], 'Inactive')
    founddevices.append(device_diff, ignore_index=True)

    writer = pd.ExcelWriter('devices.xlsx', engine='openpyxl')
    founddevices.to_excel(writer, index=False)
    device_diff.to_excel(writer, index=False, header=False, startrow=len(founddevices) + 1)
    writer.close()
    return alert

def setuptdata():
    sheet['A1'] = 'Source IP'
    sheet['B1'] = 'Destination IP'
    sheet['C1'] = 'Protocol'
    sheet['D1'] = 'Size'
    book.save('training.xlsx')


def traincapture():
    t_end = time.time() + 86400

    capture = pyshark.LiveCapture(interface='Ethernet', display_filter="ip.addr == 192.168.0.30")
    capture.sniff(timeout=1)
    capture
    while time.time() < t_end:
        for packet in capture.sniff_continuously(packet_count=10):
            row = (packet.ip.src, packet.ip.dst, packet.highest_layer, packet.captured_length)
            sheet.append(row)
            book.save('training.xlsx')

def iforest(file):
    trainingdata = pd.read_excel(file + '.xlsx')
    trainingdata.head()

    for col in trainingdata.columns:
        if trainingdata[col].dtype == "object":
            le = LabelEncoder()
            trainingdata[col].fillna("None", inplace=True)
            le.fit(list(trainingdata[col].astype(str).values))
            trainingdata[col] = le.transform(list(trainingdata[col].astype(str).values))

    model = IsolationForest(max_samples=390, contamination=float(0), n_estimators=1000)
    model.fit(trainingdata)

    trainingdata["iforest"] = pd.Series(model.predict(trainingdata))
    trainingdata["iforest"] = trainingdata["iforest"].map({1: 0, -1: 1})
    trainingdata["iforest"].value_counts()



def setupldata(runningtime):
    if runningtime == 1:
        sheet2['A1'] = 'Source IP'
        sheet2['B1'] = 'Destination IP'
        sheet2['C1'] = 'Protocol'
        sheet2['D1'] = 'Size'
        book2.save('livedata.xlsx')


def livepacketcapture():
    t_end = time.time() + 10  # 900


    capture = pyshark.LiveCapture(interface='Ethernet', display_filter="ip.addr == 192.168.0.30")
    capture.sniff(timeout=1)
    capture

    while time.time() < t_end:
        for packet in capture.sniff_continuously(packet_count=10):
            row = (packet.ip.src, packet.ip.dst, packet.highest_layer, packet.captured_length)
            sheet.append(row)
            book.save('livedata.xlsx')
    graphs()
    iforest('livedata')
    setupldata()
    livepacketcapture()


def graphfile(runningtime):
    if runningtime == 2:
        sheet3['A1'] = 'in'
        sheet3['B1'] = 'out'
        sheet3['C1'] = 'time'
        book3.save('graphdata.xlsx')

def graphmaker(time):

        df = pd.read_excel('training.xlsx')
        dataout = df[df["Source IP"] == "192.168.0.30"]['Size'].sum()
        datain = df[df["Destination IP"] == "192.168.0.30"]['Size'].sum()

        toplot = pd.read_excel('graphdata.xlsx')
        adddata = (datain, dataout, time)
        df_length = len(toplot)
        toplot.loc[df_length] = adddata

        writer = pd.ExcelWriter('graphdata.xlsx', engine='openpyxl')
        toplot.to_excel(writer, index=False)
        writer.close()






@app.route('/', methods=["GET", "POST"])
def main():
    alert = "No new devices found"
    setuphosts(runningtime)
    alert = refreshhosts(alert)

    hosts = pd.read_excel('devices.xlsx')
    return render_template('devices.html', tables=[hosts.to_html(classes='hosts', index=False)],
                           titles=['na', ''], variable=alert)


@app.route('/detection', methods=["GET", "POST"])
def detection():
    alert = 'No detections'
    if runningtime == 1:
        setuptdata()
        traincapture()
        iforest('training')
    else:
        setupldata()
        livepacketcapture()
    annomalies = pd.read_excel('annomalies.xlsx')
    return render_template('detection.html', tables=[annomalies.to_html(classes='hosts', index=False)],
                           titles=['na', ''], variable=alert)


@app.route('/graphs', methods=["GET", "POST"])
def graphs(time):
    t_end = time.time() + 15
    graphmaker(time)
    if time.time() > t_end:
        graphfile(runningtime)
        time +=15

    toplot = pd.read_excel('graphdata.xlsx')

    line_labels = toplot['time']
    line_values = toplot['in']
    line_labels1 = toplot['out']
    return render_template('graphs.html', title='Total data in and out every cycle', max=1000000, labels=line_labels,
                           values=line_values, values1=line_labels1)




if __name__ == "__main__":
    app.run(debug=True)
